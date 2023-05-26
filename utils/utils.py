from openpyxl import Workbook
from utils.settings import URL_PREFIX, OUTPUT_DATA_PATH, DETAIL_PREFIX
from utils.pivixitem import PivixItem
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

def get_start_url(urls):
    ret = []
    for u in urls:
        item = URL_PREFIX + u
        ret.append(item)
    return ret

def make_start_url(aid):
    # https://www.pixiv.net/bookmark_detail.php? illust_id=102368203&p=12
    return DETAIL_PREFIX + "?illust_id=" + aid

def make_extract_content_extent(origin:[], content:[]):
    if len(origin) == 0:
        origin.append([])
        origin.append([])

    origin[0].extend(content[0])
    origin[1].extend(content[1])
    return origin


def save_data(item:PivixItem, block):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'artwork_id'
    ws['B1'] = 'mark_users'
    ws['C1'] = 'date'
    print("-------saving data-------")
    l = len(item.followers)
    for index in range(l):
        name = item.followers[index]
        date = item.dates[index]

        legal_name = re.sub(ILLEGAL_CHARACTERS_RE, "", name)
        line = [item.aid, str(legal_name), str(date)]
        try:
            ws.append(line)
        except Exception as e:
            print(e)
    print("will begin write")
    try:
        wb.save(OUTPUT_DATA_PATH + block + '/' + f'{item.aid}.xlsx')
    except Exception as e:
        print(e)
        wb.close()
    else:
        print(item.aid, ' save √! items = ', len(item.followers))
        wb.close()

