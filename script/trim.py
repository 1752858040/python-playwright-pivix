import openpyxl
from openpyxl import Workbook
from utils.settings import ORIGIN_DATA_PATH
import math

BLOCKSIZE = 100

def trim(data, block):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'artwork_id'
    ws['B1'] = 'artist_id'
    for i in data:
        artwork = i['artwork_id']
        artist = i['artist_id']
        line = [artwork, artist]
        ws.append(line)
    try:
        wb.save(ORIGIN_DATA_PATH + f'{block}.xlsx')
    except PermissionError as e:
        print(e)
    except FileExistsError as e2:
        print(e2)
    else:
        print("block ", block)


if __name__ == '__main__':
    wb = openpyxl.load_workbook('artwork list.xlsx')
    ws = wb.active
    artwork_id = []
    artist_id = []
    index = 0
    count = 0
    c1 = ''
    c2 = ''

    data = []
    for line in ws:
        if index == 0:
            index += 1
            continue

        item = {}
        item['artwork_id'] = str(line[0].value)
        item['artist_id'] = str(line[1].value)
        data.append(item)

        index += 1
        count += 1

        if count % BLOCKSIZE == 0:
            block = math.floor(count / BLOCKSIZE)
            print('block  save', block)
            trim(data, block)
            data.clear()

